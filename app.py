import os
import io
import smtplib
from datetime import date, datetime
import time
import hashlib
import random
import re
from urllib.parse import urlparse
from functools import wraps
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask import Flask, render_template, request, jsonify, redirect, url_for, session, send_file
from twilio.rest import Client
import mysql.connector
from mysql.connector import pooling
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "school_secret_key_2024")

print("=" * 50)
print("Flask starting...")
print("DB_URL        :", "SET" if os.getenv("DB_URL") else "MISSING")
print("Twilio SID    :", "SET" if os.getenv("TWILIO_ACCOUNT_SID") else "MISSING")
print("Twilio SMS NUM:", os.getenv("TWILIO_NUMBER") or "MISSING")
print("Gmail User    :", os.getenv("GMAIL_USER") or "MISSING")
print("=" * 50)

# ================= CONFIG =================
ADMIN_USERNAME         = os.getenv("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD         = os.getenv("ADMIN_PASSWORD", "school@123")
SCHOOL_NAME            = os.getenv("SCHOOL_NAME", "Our School")
SCHOOL_CODE            = os.getenv("SCHOOL_CODE", "SCH")
GMAIL_USER             = os.getenv("GMAIL_USER", "")
GMAIL_PASSWORD         = os.getenv("GMAIL_PASSWORD", "")
TWILIO_NUMBER          = os.getenv("TWILIO_NUMBER", "")
TWILIO_WHATSAPP_NUMBER = os.getenv("TWILIO_WHATSAPP_NUMBER", "")

# ================= FILE UPLOAD =================
UPLOAD_FOLDER = "uploads"
ALLOWED_PDF   = {"pdf"}
ALLOWED_IMG   = {"jpg", "jpeg", "png"}
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def allowed_pdf(f): return "." in f and f.rsplit(".",1)[1].lower() in ALLOWED_PDF
def allowed_img(f): return "." in f and f.rsplit(".",1)[1].lower() in ALLOWED_IMG

# ================= DATABASE =================
DB_URL = os.getenv("DB_URL")
if not DB_URL:
    raise Exception("DB_URL not found in .env")

_url = urlparse(DB_URL)
db_pool = pooling.MySQLConnectionPool(
    pool_name="student_pool", pool_size=5,
    host=_url.hostname, user=_url.username,
    password=_url.password, database=_url.path.lstrip("/"),
    port=_url.port or 3306
)

def get_db(): return db_pool.get_connection()

def init_db():
    try:
        conn   = get_db()
        cursor = conn.cursor()
        for s in [
            "ALTER TABLE students ADD COLUMN reg_no VARCHAR(20)",
            "ALTER TABLE students ADD COLUMN email VARCHAR(150)",
            "ALTER TABLE students ADD COLUMN photo VARCHAR(255)",
        ]:
            try: cursor.execute(s)
            except: pass
        # Create settings table for deadline
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                setting_key   VARCHAR(100) PRIMARY KEY,
                setting_value VARCHAR(255)
            )
        """)
        conn.commit()
        cursor.close(); conn.close()
        print("[DB] Init done ✅")
    except Exception as e:
        print(f"[DB] init warning: {e}")

init_db()

# ================= TWILIO =================
ACCOUNT_SID   = os.getenv("TWILIO_ACCOUNT_SID")
AUTH_TOKEN    = os.getenv("TWILIO_AUTH_TOKEN")
twilio_client = Client(ACCOUNT_SID, AUTH_TOKEN)

otp_store    = {}
otp_verified = set()
OTP_EXPIRY   = 300

# ================= HELPERS =================
def normalize_phone(phone):
    phone = phone.strip() if phone else ""
    phone = re.sub(r"^\+91", "", phone).strip()
    return "+91" + phone

def hash_aadhaar(a): return hashlib.sha256(a.encode()).hexdigest()
def valid_aadhaar(a): return bool(re.fullmatch(r"\d{12}", a))
def valid_pan(p): return bool(re.fullmatch(r"[A-Z]{5}[0-9]{4}[A-Z]", p.upper()))
def valid_mobile(p): return bool(re.fullmatch(r"\+91\d{10}", p))

def generate_reg_no():
    year   = date.today().year
    conn   = get_db(); cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM students WHERE reg_no IS NOT NULL")
    count  = cursor.fetchone()[0] + 1
    cursor.close(); conn.close()
    return f"{SCHOOL_CODE}-{year}-{str(count).zfill(4)}"

def save_pdf(file):
    if file and file.filename and allowed_pdf(file.filename):
        fn = secure_filename(file.filename)
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], fn))
        return fn
    return None

def save_photo(file):
    if file and file.filename and allowed_img(file.filename):
        ext = file.filename.rsplit(".",1)[1].lower()
        fn  = f"photo_{int(time.time())}_{random.randint(1000,9999)}.{ext}"
        file.save(os.path.join(app.config["UPLOAD_FOLDER"], fn))
        return fn
    return None

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("admin_logged_in"):
            return redirect(url_for("admin_login"))
        return f(*args, **kwargs)
    return decorated

# ================= DEADLINE HELPER =================
def get_deadline():
    """Returns deadline date string or None."""
    try:
        conn   = get_db(); cursor = conn.cursor()
        cursor.execute("SELECT setting_value FROM settings WHERE setting_key='deadline'")
        row = cursor.fetchone()
        cursor.close(); conn.close()
        return row[0] if row else None
    except: return None

def is_deadline_passed():
    dl = get_deadline()
    if not dl: return False
    try:
        return date.today() > datetime.strptime(dl, "%Y-%m-%d").date()
    except: return False

# ================= SEAT AVAILABILITY =================
def get_seat_info(class_name):
    """Returns (total, filled, available) for a class."""
    try:
        conn   = get_db(); cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT total_seats, filled_seats FROM classes WHERE class_name=%s", (class_name,))
        row = cursor.fetchone()
        cursor.close(); conn.close()
        if row:
            avail = row["total_seats"] - row["filled_seats"]
            return row["total_seats"], row["filled_seats"], avail
    except: pass
    return None, None, None

def get_all_seat_info():
    """Returns list of all classes with seat info."""
    try:
        conn   = get_db(); cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT class_name, total_seats, filled_seats FROM classes ORDER BY id")
        rows = cursor.fetchall()
        cursor.close(); conn.close()
        for r in rows:
            r["available"] = r["total_seats"] - r["filled_seats"]
            r["pct"]       = round((r["filled_seats"] / r["total_seats"]) * 100) if r["total_seats"] else 0
        return rows
    except: return []

# ================================================================
# =================== NOTIFICATION FUNCTIONS =====================
# ================================================================
def send_sms(to, msg):
    if not TWILIO_NUMBER: return
    try:
        twilio_client.messages.create(body=msg, from_=TWILIO_NUMBER, to=to)
        print(f"[SMS] Sent to {to}")
    except Exception as e: print(f"[SMS] ERROR: {e}")

def send_whatsapp(to, msg):
    if not TWILIO_WHATSAPP_NUMBER: return
    try:
        twilio_client.messages.create(body=msg, from_=TWILIO_WHATSAPP_NUMBER, to=f"whatsapp:{to}")
        print(f"[WA] Sent to {to}")
    except Exception as e: print(f"[WA] ERROR: {e}")

def send_email(to, subject, html):
    if not GMAIL_USER or not GMAIL_PASSWORD or not to or "@" not in to: return
    try:
        m = MIMEMultipart("alternative")
        m["Subject"] = subject; m["From"] = f"{SCHOOL_NAME} <{GMAIL_USER}>"; m["To"] = to
        m.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_USER, GMAIL_PASSWORD); s.sendmail(GMAIL_USER, to, m.as_string())
        print(f"[Email] Sent to {to}")
    except Exception as e: print(f"[Email] ERROR: {e}")

def notify_registration(name, phone, email, cls, reg_no):
    sms = f"Dear Parent,\nRegistration successful!\nStudent: {name}\nClass: {cls}\nReg No: {reg_no}\nWe will notify you once reviewed.\n- {SCHOOL_NAME}"
    wa  = f"🏫 *{SCHOOL_NAME}*\n\nDear Parent,\n\n✅ *Registration Successful!*\n\n*Student:* {name}\n*Class:* {cls}\n*Reg No:* `{reg_no}`\n\nSave your Reg No for future reference.\n\n_- {SCHOOL_NAME} Administration_"
    html = f"<div style='font-family:Arial;max-width:600px;margin:auto;'><div style='background:#1a3c6e;padding:24px;text-align:center;'><h1 style='color:white;'>🏫 {SCHOOL_NAME}</h1></div><div style='padding:32px;'><h2 style='color:#1a3c6e;'>Registration Received ✅</h2><p>Dear Parent, <strong>{name}'s</strong> application for <strong>{cls}</strong> has been received.</p><div style='background:#f0f4f8;border-radius:8px;padding:20px;margin:20px 0;text-align:center;'><p style='color:#888;font-size:13px;'>Registration Number</p><p style='font-size:28px;font-weight:700;color:#1a3c6e;letter-spacing:2px;'>{reg_no}</p></div><p style='color:#999;font-size:13px;'>- {SCHOOL_NAME}</p></div></div>"
    send_sms(phone, sms); send_whatsapp(phone, wa); send_email(email, f"Registration Confirmed [{reg_no}]", html)

def notify_accepted(name, phone, email, cls, fees_date, fees_time, reg_no):
    sms = f"Dear Parent,\nCongratulations! {name} (Reg: {reg_no}) for {cls} at {SCHOOL_NAME} is ACCEPTED.\nVisit on {fees_date} at {fees_time} for fees.\nBring all original documents.\n- {SCHOOL_NAME}"
    wa  = f"🏫 *{SCHOOL_NAME}*\n\nDear Parent,\n\n🎉 *Admission Accepted!*\n\n*Student:* {name}\n*Reg No:* `{reg_no}`\n*Class:* {cls}\n\n📅 *Fees Date:* {fees_date}\n🕐 *Time:* {fees_time}\n\n📋 Bring original documents, birth certificate, 4 photos.\n\n_- {SCHOOL_NAME} Administration_"
    html = f"<div style='font-family:Arial;max-width:600px;margin:auto;'><div style='background:#1a9e5c;padding:24px;text-align:center;'><h1 style='color:white;'>🎉 Admission Accepted!</h1><p style='color:#a0e0c0;'>{SCHOOL_NAME}</p></div><div style='padding:32px;'><p>Dear Parent, <strong>{name}'s</strong> admission for <strong>{cls}</strong> has been <strong>ACCEPTED</strong>.</p><div style='background:#e8f5e9;border:1.5px solid #a5d6a7;border-radius:8px;padding:20px;margin:20px 0;text-align:center;'><p style='font-size:13px;color:#555;'>Please visit for fees submission</p><p style='font-size:26px;font-weight:700;color:#1a9e5c;'>📅 {fees_date}</p><p>🕐 {fees_time}</p></div><p style='color:#999;font-size:13px;'>- {SCHOOL_NAME}</p></div></div>"
    send_sms(phone, sms); send_whatsapp(phone, wa); send_email(email, f"Admission Accepted [{reg_no}]", html)

def notify_rejected(name, phone, email, cls, reg_no=""):
    sms = f"Dear Parent,\nWe regret {name}'s admission for {cls} at {SCHOOL_NAME} could not be accepted.\nContact school for details.\n- {SCHOOL_NAME}"
    wa  = f"🏫 *{SCHOOL_NAME}*\n\nDear Parent,\n\nWe regret that *{name}'s* application for *{cls}* has not been accepted.\nPlease contact the school for more information.\n\n_- {SCHOOL_NAME} Administration_"
    html = f"<div style='font-family:Arial;max-width:600px;margin:auto;'><div style='background:#d63031;padding:24px;text-align:center;'><h1 style='color:white;'>Application Update</h1></div><div style='padding:32px;'><p>Dear Parent, we regret that <strong>{name}'s</strong> application for <strong>{cls}</strong> has not been accepted at this time. Please contact school for details.</p><p style='color:#999;font-size:13px;'>- {SCHOOL_NAME}</p></div></div>"
    send_sms(phone, sms); send_whatsapp(phone, wa); send_email(email, f"Application Update - {name}", html)

# ================================================================
# ======================== ROUTES ================================
# ================================================================

@app.route("/")
def home():
    deadline   = get_deadline()
    closed     = is_deadline_passed()
    seat_info  = get_all_seat_info()
    return render_template("form.html", deadline=deadline, closed=closed, seat_info=seat_info)

# -------- SEAT AVAILABILITY API --------
@app.route("/seat_info")
def seat_info():
    return jsonify(get_all_seat_info())

# -------- SEND OTP --------
@app.route("/send_otp", methods=["POST"])
def send_otp():
    if is_deadline_passed():
        return jsonify({"status": "error", "message": "Registration deadline has passed."})
    raw   = request.form.get("phone", "")
    phone = normalize_phone(raw)
    if not valid_mobile(phone):
        return jsonify({"status": "error", "message": "Invalid mobile number"})
    otp = str(random.randint(100000, 999999))
    otp_store[phone] = {"otp": otp, "time": time.time()}
    print(f"[OTP] {otp} → {phone}")
    try:
        msg = twilio_client.messages.create(
            body=f"Your {SCHOOL_NAME} registration OTP is {otp}. Valid 5 mins.",
            from_=TWILIO_NUMBER, to=phone)
        print(f"[OTP] SID: {msg.sid}")
        return jsonify({"status": "success"})
    except Exception as e:
        print(f"[OTP] ERROR: {e}")
        return jsonify({"status": "error", "message": str(e)})

# -------- VERIFY OTP --------
@app.route("/verify_otp", methods=["POST"])
def verify_otp():
    phone = normalize_phone(request.form.get("phone", ""))
    otp   = request.form.get("otp", "").strip()
    data  = otp_store.get(phone)
    if not data:
        return jsonify({"status": "error", "message": "No OTP. Please send OTP first."})
    if time.time() - data["time"] > OTP_EXPIRY:
        otp_store.pop(phone, None)
        return jsonify({"status": "error", "message": "OTP expired. Resend."})
    if otp == data["otp"]:
        otp_verified.add(phone); otp_store.pop(phone, None)
        return jsonify({"status": "success"})
    return jsonify({"status": "error", "message": "Incorrect OTP."})

# -------- SUBMIT FORM --------
@app.route("/submit", methods=["POST"])
def submit():
    if is_deadline_passed():
        return "Registration deadline has passed.", 400
    phone = normalize_phone(request.form.get("mobile", ""))
    if phone not in otp_verified:
        return "Mobile not verified.", 400

    name                 = request.form.get("name","").strip()
    father_name          = request.form.get("father_name","").strip()
    date_of_birth        = request.form.get("date_of_birth","").strip()
    address              = request.form.get("address","").strip()
    father_occupation    = request.form.get("father_occupation","").strip()
    academic_year        = request.form.get("academic_year","").strip()
    previous_institution = request.form.get("previous_institution_name","").strip()
    class_applied        = request.form.get("class_applied","").strip()
    category             = request.form.get("category","").strip()
    gender               = request.form.get("gender","").strip()
    special_child        = request.form.get("special_child","no")
    extra_activity       = request.form.get("extra_activity","no")
    achievement          = request.form.get("achievement","no")
    hobbies              = request.form.get("hobbies","").strip()
    sports               = request.form.get("sports","").strip()
    aadhaar              = request.form.get("aadhaar","").strip()
    pan_no               = request.form.get("pan_no","").strip().upper()
    email                = request.form.get("email","").strip().lower()

    errors = []
    if not name:                   errors.append("Name is required")
    if not father_name:            errors.append("Father name is required")
    if not date_of_birth:          errors.append("DOB is required")
    if not class_applied:          errors.append("Class is required")
    if not category:               errors.append("Category is required")
    if not valid_aadhaar(aadhaar): errors.append("Invalid Aadhaar")
    if not valid_pan(pan_no):      errors.append("Invalid PAN")
    if errors: return "<br>".join(errors), 400

    # Check seat availability
    total, filled, avail = get_seat_info(class_applied)
    if total is not None and avail <= 0:
        return f"Sorry, no seats available for {class_applied}. All {total} seats are filled.", 400

    reg_no           = generate_reg_no()
    aadhaar_hash     = hash_aadhaar(aadhaar)
    photo            = save_photo(request.files.get("photo"))
    special_file     = save_pdf(request.files.get("special_file"))
    extra_file       = save_pdf(request.files.get("extra_file"))
    achievement_file = save_pdf(request.files.get("achievement_file"))

    conn = get_db(); cursor = conn.cursor()
    try:
        cursor.execute("""
            INSERT INTO students (
                reg_no, name, father_name, date_of_birth, address,
                father_occupation, academic_year, previous_institution_name,
                class_applied, category, gender,
                phone_no, aadhaar_no, pan_no, email, photo,
                special_child, extra_activity, achievement,
                hobbies, sports,
                special_file, extra_file, achievement_file, status
            ) VALUES (
                %s,%s,%s,%s,%s, %s,%s,%s, %s,%s,%s,
                %s,%s,%s,%s,%s, %s,%s,%s, %s,%s, %s,%s,%s,%s
            )
        """, (
            reg_no, name, father_name, date_of_birth, address,
            father_occupation, academic_year, previous_institution,
            class_applied, category, gender,
            phone, aadhaar_hash, pan_no, email, photo,
            special_child, extra_activity, achievement,
            hobbies, sports,
            special_file, extra_file, achievement_file, "pending"
        ))
        conn.commit()
        otp_verified.discard(phone)
        print(f"[SUBMIT] {name} | {reg_no}")
        notify_registration(name, phone, email, class_applied, reg_no)
    except Exception as e:
        conn.rollback(); print(f"[SUBMIT] DB ERROR: {e}")
        return f"Registration failed: {e}", 500
    finally:
        cursor.close(); conn.close()

    return redirect(url_for("success", reg=reg_no, student=name))

# -------- SUCCESS --------
@app.route("/success")
def success():
    return render_template("success.html",
                           reg_no=request.args.get("reg",""),
                           student_name=request.args.get("student",""))

# -------- PUBLIC STATUS CHECK --------
@app.route("/students")
def students():
    return render_template("students.html")

@app.route("/check_status", methods=["POST"])
def check_status():
    query = request.form.get("query","").strip()
    phone = normalize_phone(query) if query.isdigit() and len(query)==10 else None
    conn  = get_db(); cursor = conn.cursor(dictionary=True)
    if phone:
        cursor.execute("SELECT reg_no,name,father_name,class_applied,category,gender,phone_no,status FROM students WHERE phone_no=%s ORDER BY id DESC LIMIT 1",(phone,))
    else:
        cursor.execute("SELECT reg_no,name,father_name,class_applied,category,gender,phone_no,status FROM students WHERE reg_no=%s ORDER BY id DESC LIMIT 1",(query.upper(),))
    s = cursor.fetchone(); cursor.close(); conn.close()
    return jsonify({"found": True, "student": s} if s else {"found": False})

# -------- ADMIN LOGIN --------
@app.route("/admin/login", methods=["GET","POST"])
def admin_login():
    if request.method == "POST":
        if request.form.get("username")==ADMIN_USERNAME and request.form.get("password")==ADMIN_PASSWORD:
            session["admin_logged_in"] = True
            return redirect(url_for("admin_panel"))
        return render_template("admin_login.html", error="Invalid credentials")
    return render_template("admin_login.html", error=None)

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin_logged_in", None)
    return redirect(url_for("admin_login"))

# -------- ADMIN PANEL --------
@app.route("/admin")
@admin_required
def admin_panel():
    search         = request.args.get("search","").strip()
    filter_class   = request.args.get("class_applied","").strip()
    filter_status  = request.args.get("status","").strip()

    conn = get_db(); cursor = conn.cursor(dictionary=True)
    q = "SELECT id,reg_no,name,father_name,class_applied,category,gender,phone_no,email,photo,status FROM students WHERE 1=1"
    params = []
    if search:
        q += " AND (name LIKE %s OR reg_no LIKE %s OR phone_no LIKE %s)"
        params += [f"%{search}%",f"%{search}%",f"%{search}%"]
    if filter_class:  q += " AND class_applied=%s"; params.append(filter_class)
    if filter_status: q += " AND status=%s";        params.append(filter_status)
    q += " ORDER BY id DESC"
    cursor.execute(q, params)
    data       = cursor.fetchall()
    seat_info  = get_all_seat_info()
    deadline   = get_deadline()
    cursor.close(); conn.close()
    return render_template("admin.html", students=data, seat_info=seat_info,
                           deadline=deadline, search=search,
                           filter_class=filter_class, filter_status=filter_status)

# -------- SET DEADLINE (ADMIN) --------
@app.route("/admin/set_deadline", methods=["POST"])
@admin_required
def set_deadline():
    dl = request.form.get("deadline","").strip()
    conn = get_db(); cursor = conn.cursor()
    if dl:
        cursor.execute("REPLACE INTO settings (setting_key, setting_value) VALUES ('deadline', %s)", (dl,))
    else:
        cursor.execute("DELETE FROM settings WHERE setting_key='deadline'")
    conn.commit(); cursor.close(); conn.close()
    return redirect(url_for("admin_panel"))

# -------- UPDATE SEATS (ADMIN) --------
@app.route("/admin/update_seats", methods=["POST"])
@admin_required
def update_seats():
    class_name  = request.form.get("class_name","").strip()
    total_seats = request.form.get("total_seats","0").strip()
    conn = get_db(); cursor = conn.cursor()
    cursor.execute("UPDATE classes SET total_seats=%s WHERE class_name=%s", (int(total_seats), class_name))
    conn.commit(); cursor.close(); conn.close()
    return redirect(url_for("admin_panel"))

# -------- EXPORT TO EXCEL --------
@app.route("/admin/export_excel")
@admin_required
def export_excel():
    filter_class  = request.args.get("class_applied","").strip()
    filter_status = request.args.get("status","").strip()

    conn = get_db(); cursor = conn.cursor(dictionary=True)
    q = """SELECT reg_no, name, father_name, date_of_birth, gender, category,
                  class_applied, phone_no, email, father_occupation,
                  academic_year, previous_institution_name,
                  special_child, extra_activity, achievement,
                  hobbies, sports, status, created_at
           FROM students WHERE 1=1"""
    params = []
    if filter_class:  q += " AND class_applied=%s"; params.append(filter_class)
    if filter_status: q += " AND status=%s";        params.append(filter_status)
    q += " ORDER BY id"
    cursor.execute(q, params)
    rows = cursor.fetchall(); cursor.close(); conn.close()

    # ===== Build Excel =====
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Registrations"

    # Header style
    header_fill   = PatternFill("solid", fgColor="1A3C6E")
    header_font   = Font(color="FFFFFF", bold=True, size=11)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Title row
    ws.merge_cells("A1:S1")
    title_cell = ws["A1"]
    title_cell.value    = f"{SCHOOL_NAME} — Student Registrations ({date.today().strftime('%d %b %Y')})"
    title_cell.font     = Font(bold=True, size=14, color="1A3C6E")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = [
        "Reg No", "Student Name", "Father Name", "Date of Birth",
        "Gender", "Category", "Class Applied", "Mobile", "Email",
        "Father Occupation", "Acad. Year", "Previous School",
        "Special Child", "Extra Activity", "Achievement",
        "Hobbies", "Sports", "Status", "Applied On"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = center_align
        cell.border    = thin_border
    ws.row_dimensions[2].height = 22

    # Status fill colours
    status_fills = {
        "accepted": PatternFill("solid", fgColor="C8E6C9"),
        "rejected": PatternFill("solid", fgColor="FFCDD2"),
        "pending":  PatternFill("solid", fgColor="FFF9C4"),
    }

    # Data rows
    for r, row in enumerate(rows, 3):
        values = [
            row.get("reg_no",""), row.get("name",""), row.get("father_name",""),
            str(row.get("date_of_birth","")), row.get("gender",""), row.get("category",""),
            row.get("class_applied",""), row.get("phone_no",""), row.get("email",""),
            row.get("father_occupation",""), row.get("academic_year",""),
            row.get("previous_institution_name",""),
            row.get("special_child",""), row.get("extra_activity",""), row.get("achievement",""),
            row.get("hobbies",""), row.get("sports",""), row.get("status",""),
            str(row.get("created_at",""))[:10]
        ]
        status = row.get("status","pending")
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=col, value=val)
            cell.border    = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if status in status_fills:
                cell.fill = status_fills[status]

    # Column widths
    widths = [14,20,20,14,10,12,13,14,24,18,12,22,13,14,12,16,16,12,14]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Freeze header rows
    ws.freeze_panes = "A3"

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)

    fname = f"Students_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# -------- UPLOADS --------
@app.route("/uploads/<filename>")
@admin_required
def uploaded_file(filename):
    from flask import send_from_directory
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename)

# -------- APPROVE --------
@app.route("/approve/<int:student_id>", methods=["GET","POST"])
@admin_required
def approve(student_id):
    if request.method == "GET":
        conn = get_db(); cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT id,reg_no,name,class_applied,phone_no,email FROM students WHERE id=%s",(student_id,))
        student = cursor.fetchone(); cursor.close(); conn.close()
        if not student: return "Student not found", 404
        return render_template("approve_confirm.html", student=student, today=date.today().isoformat())

    fees_date = request.form.get("fees_date","").strip()
    fees_time = request.form.get("fees_time","9:00 AM").strip()
    conn = get_db(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT * FROM students WHERE id=%s",(student_id,))
    student = cursor.fetchone()
    if not student: cursor.close(); conn.close(); return "Not found", 404

    cls    = student["class_applied"]
    phone  = student["phone_no"]
    name   = student["name"]
    email  = student.get("email","")
    reg_no = student.get("reg_no","")

    cursor.execute("SELECT * FROM classes WHERE class_name=%s",(cls,))
    c = cursor.fetchone()
    if c and c["filled_seats"] < c["total_seats"]:
        cursor.execute("UPDATE students SET status='accepted' WHERE id=%s",(student_id,))
        cursor.execute("UPDATE classes SET filled_seats=filled_seats+1 WHERE class_name=%s",(cls,))
        conn.commit()
        notify_accepted(name, phone, email, cls, fees_date, fees_time, reg_no)
    else:
        cursor.execute("UPDATE students SET status='rejected' WHERE id=%s",(student_id,))
        conn.commit()
        notify_rejected(name, phone, email, cls, reg_no)
    cursor.close(); conn.close()
    return redirect(url_for("admin_panel"))

# -------- REJECT --------
@app.route("/reject/<int:student_id>")
@admin_required
def reject(student_id):
    conn = get_db(); cursor = conn.cursor(dictionary=True)
    cursor.execute("SELECT name,reg_no,class_applied,phone_no,email FROM students WHERE id=%s",(student_id,))
    s = cursor.fetchone()
    if s:
        cursor.execute("UPDATE students SET status='rejected' WHERE id=%s",(student_id,))
        conn.commit()
        notify_rejected(s["name"],s["phone_no"],s.get("email",""),s["class_applied"],s.get("reg_no",""))
    cursor.close(); conn.close()
    return redirect(url_for("admin_panel"))

# -------- ERROR HANDLERS --------
@app.errorhandler(405)
def method_not_allowed(e): return redirect(url_for("home"))
@app.errorhandler(404)
def not_found(e): return redirect(url_for("home"))

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=True)
